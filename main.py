import re
import openpyxl
from datetime import datetime

from analisis_apellido import analizar_apellido
from analisis_calificaciones import analizar_calificacion
from analisis_cedula import analizar_cedula
from analisis_correo import analizar_correo
from analisis_direcciones import analizar_direccion
from analisis_edad import analizar_edad
from analisis_fecha_de_nacimiento import analizar_fecha_de_nacimiento
from analisis_nombres import analizar_nombre
from analisis_telefono import analizar_telefono

nombre_archivo = "2020A_calidad_datos.D.J.xlsx"
excel_dataframe = openpyxl.load_workbook(nombre_archivo)
data = []

def extraer_datos_de_archivo_excel():
    dataframe = excel_dataframe.active
    for row in range(1, dataframe.max_row):
        _row = [row,]
        for col in dataframe.iter_cols(1, dataframe.max_column):
            _row.append(col[row].value)
        data.append(_row)

if __name__ == "__main__":
    extraer_datos_de_archivo_excel()
    analizar_cedula(data,nombre_archivo)
    analizar_direccion(data,nombre_archivo)
    analizar_telefono(data,nombre_archivo)
    analizar_fecha_de_nacimiento(data,nombre_archivo)
    analizar_edad(data,nombre_archivo)
    analizar_calificacion(data,nombre_archivo)
    analizar_nombre(data,nombre_archivo)
    analizar_apellido(data,nombre_archivo)
    analizar_correo(data, nombre_archivo)