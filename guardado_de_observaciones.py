import openpyxl

def guardar_observaciones_nuevo_archivo(tabla_datos_erroneos, nombre):
    libro_excel = openpyxl.Workbook()
    hoja_activa = libro_excel.active

    for fila_datos in tabla_datos_erroneos:
        hoja_activa.append(fila_datos)

    libro_excel.save(nombre + '.xlsx')

def guadar_observaciones_nueva_hoja(tabla_datos_erroneos, nombre,archivo_existente):
    libro_excel_existente = openpyxl.load_workbook(archivo_existente)
    nueva_hoja = libro_excel_existente.create_sheet(nombre)

    # Llenar la nueva hoja con datos desde la lista
    for fila_datos in tabla_datos_erroneos:
        nueva_hoja.append(fila_datos)

    libro_excel_existente.save(archivo_existente)